import os
import csv
import PyPDF4
from pathlib import Path
from datetime import datetime
import re
from typing import List, Dict, Tuple
import shutil
from tqdm import tqdm
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_TKINTERDND = True
except ImportError:
    HAS_TKINTERDND = False
import subprocess
import sys
import argparse
import platform
from docx import Document
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import black
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PyPDF2 import PdfMerger
import xlrd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from PIL import Image
import email
from email import policy
from email.parser import BytesParser
from fpdf import FPDF
import html

# Windows-specific imports
if platform.system() == 'Windows':
    import win32com.client
    import comtypes.client

class EnhancedBatesNumbering:
    def __init__(self, input_dir: str, output_dir: str, prefix: str = '', 
                 zero_pad_length: int = 6, start: int = 1, is_single_file: bool = False,
                 stamp_x: float = 0.97, stamp_y: float = 0.001, stamp_color: str = "black", 
                 stamp_box_width: float = 0.0, stamp_position: str = "bottom-right",
                 stamp_x_offset: int = 0, stamp_y_offset: int = 0, stamp_opacity: int = 100):
        self.input_dir = Path(input_dir)
        self.is_single_file = is_single_file
        
        if is_single_file:
            # For single files, use the output directory directly
            self.output_dir = Path(output_dir)
        else:
            # For directories, create timestamped output directory
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.output_dir = Path(output_dir) / f"BATES_{prefix}_{timestamp}"
        
        self.prefix = prefix
        self.zero_pad_length = zero_pad_length
        self.start = start
        self.current_number = start
        
        # Stamp appearance settings
        self.stamp_x = stamp_x
        self.stamp_y = stamp_y
        self.stamp_color = stamp_color
        self.stamp_box_width = stamp_box_width
        self.stamp_position = stamp_position
        self.stamp_x_offset = stamp_x_offset
        self.stamp_y_offset = stamp_y_offset
        self.stamp_opacity = stamp_opacity / 100.0  # Convert to 0-1 range
        
        # Only create issues directory for directory processing
        if not is_single_file:
            self.issues_dir = self.output_dir / "_FILES WITH ISSUES"
            self.issues_dir.mkdir(parents=True, exist_ok=True)
        
        # Define ignored file types and patterns
        self.ignored_files = {
            '.DS_Store',  # macOS system files
            '.Thumbs.db',  # Windows thumbnail files
            '~$',  # Temporary Office files
            '.tmp',  # Temporary files
            '.temp',  # Temporary files
            '.bak',  # Backup files
            '.swp',  # Vim swap files
            '.swo',  # Vim swap files
            '.log',  # Log files
            '.ini',  # Configuration files
            '.db',  # Database files
            '.sqlite',  # SQLite database files
            '.sqlite3',  # SQLite database files
            '.db-shm',  # SQLite shared memory files
            '.db-wal',  # SQLite write-ahead log files
        }
        
        # Define supported spreadsheet types
        self.spreadsheet_types = {
            '.csv', '.xls', '.xlsx', '.xlsm', '.xlsb', '.ods', '.numbers'
        }
        
        self.setup_logging()
        
        # Check for required dependencies
        self.check_dependencies()
        
    def check_dependencies(self):
        """Check and install required dependencies."""
        try:
            import pycryptodome
            self.has_pycryptodome = True
        except ImportError:
            self.logger.info("Installing pycryptodome for PDF encryption support...")
            try:
                subprocess.run(['pip', 'install', 'pycryptodome'], check=True)
                self.has_pycryptodome = True
            except Exception as e:
                self.logger.warning(f"Failed to install pycryptodome: {str(e)}")
                self.has_pycryptodome = False

    def setup_logging(self):
        # Create output directory first
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Only create log file if not processing a single file
        if not self.is_single_file:
            # Create log filename with prefix and timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_filename = f"BATES_{self.prefix}_{timestamp}_processing.log"
            log_file = self.output_dir / log_filename
            
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_file),
                    logging.StreamHandler()
                ]
            )
            self.logger = logging.getLogger(__name__)
            self.logger.info(f"Logging initialized with prefix: {self.prefix}")
        else:
            # For single file processing, just use basic console logging
            logging.basicConfig(
                level=logging.INFO,
                format='%(message)s',
                handlers=[logging.StreamHandler()]
            )
            self.logger = logging.getLogger(__name__)

    def extract_date_from_pdf(self, pdf_path: Path) -> Tuple[datetime, datetime]:
        """Extract the first date found in the PDF content and return both extracted and creation dates."""
        creation_date = datetime.fromtimestamp(pdf_path.stat().st_ctime)
        extracted_date = None
        
        try:
            # First try to get metadata date
            try:
                with open(pdf_path, 'rb') as file:
                    reader = PyPDF4.PdfFileReader(file)
                    if reader.getDocumentInfo():
                        # Try various metadata fields that might contain dates
                        metadata = reader.getDocumentInfo()
                        date_fields = [
                            '/CreationDate',
                            '/ModDate',
                            '/Date',
                            '/LastModified',
                            '/LastPrinted'
                        ]
                        
                        for field in date_fields:
                            if field in metadata:
                                date_str = metadata[field]
                                # Clean up PDF date format (e.g., "D:20231115123456-06'00'")
                                date_str = re.sub(r"D:|'00'", "", date_str)
                                try:
                                    # Try parsing the date string
                                    extracted_date = datetime.strptime(date_str[:8], '%Y%m%d')
                                    if extracted_date and extracted_date <= datetime.now():
                                        self.logger.info(f"Found date in metadata for {pdf_path}: {extracted_date}")
                                        return extracted_date, creation_date
                                except ValueError:
                                    continue
            except Exception as e:
                self.logger.debug(f"Could not read metadata from {pdf_path}: {str(e)}")
            
            # If no metadata date, try content extraction
            try:
                with open(pdf_path, 'rb') as file:
                    reader = PyPDF4.PdfFileReader(file)
                    # Look through first few pages for dates
                    for page_num in range(min(3, reader.getNumPages())):
                        try:
                            page = reader.getPage(page_num)
                            text = page.extractText()
                            
                            # Common date patterns
                            date_patterns = [
                                # Standard numeric formats
                                r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}',  # MM/DD/YYYY
                                r'\d{4}[-/]\d{1,2}[-/]\d{1,2}',    # YYYY/MM/DD
                                
                                # Month names and abbreviations
                                r'\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{2,4}',
                                r'\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4}',
                                
                                # Month names with optional day
                                r'(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{2,4}',
                                r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{2,4}',
                                
                                # ISO format
                                r'\d{4}-\d{2}-\d{2}',
                                
                                # European format
                                r'\d{1,2}\.\d{1,2}\.\d{2,4}',
                                
                                # Military format
                                r'\d{2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4}',
                                
                                # With optional time
                                r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?',
                                
                                # With optional timezone
                                r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?\s*(?:EST|CST|MST|PST|UTC|GMT)?',
                                
                                # Additional patterns
                                r'\d{4}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2}',
                                r'\d{4}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2}',
                                r'(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}',
                                r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{4}'
                            ]
                            
                            for pattern in date_patterns:
                                matches = re.findall(pattern, text, re.IGNORECASE)
                                if matches:
                                    try:
                                        # Clean up the date string
                                        date_str = matches[0].strip()
                                        
                                        # Handle various separators
                                        date_str = re.sub(r'[/.]', '-', date_str)
                                        
                                        # Handle month names
                                        month_map = {
                                            'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
                                            'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
                                            'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
                                        }
                                        
                                        # Try to parse the date string
                                        try:
                                            # First try direct parsing
                                            extracted_date = datetime.strptime(date_str.split()[0], '%Y-%m-%d')
                                        except ValueError:
                                            # Try parsing with month names
                                            parts = date_str.split()
                                            if len(parts) >= 3:
                                                month = parts[1][:3].lower()
                                                if month in month_map:
                                                    # Reconstruct date string with numeric month
                                                    date_str = f"{parts[2]}-{month_map[month]}-{parts[0].zfill(2)}"
                                                    extracted_date = datetime.strptime(date_str, '%Y-%m-%d')
                                        
                                        if extracted_date:
                                            # Validate reasonable date range (e.g., not in the future)
                                            if extracted_date > datetime.now():
                                                self.logger.warning(f"Found future date in {pdf_path}: {extracted_date}")
                                                continue
                                            self.logger.info(f"Found date in content for {pdf_path}: {extracted_date}")
                                            return extracted_date, creation_date
                                            
                                    except ValueError as e:
                                        self.logger.debug(f"Could not parse date '{date_str}' from {pdf_path}: {str(e)}")
                                        continue
                        except Exception as e:
                            self.logger.debug(f"Error processing page {page_num} of {pdf_path}: {str(e)}")
                            continue
                            
            except Exception as e:
                self.logger.warning(f"Could not extract date from {pdf_path}: {str(e)}")
                
        except Exception as e:
            self.logger.warning(f"Error processing {pdf_path}: {str(e)}")
        
        # If we get here, we couldn't find a valid date
        self.logger.info(f"No valid date found in {pdf_path}, using creation date: {creation_date}")
        return None, creation_date

    def generate_description(self, file_path: Path) -> str:
        """Generate description based on folder and file names."""
        folder_name = file_path.parent.name
        file_name = file_path.stem
        return f"{folder_name} - {file_name}"

    def convert_to_pdf(self, input_path: Path, output_dir: Path) -> Path:
        """Convert non-PDF files to PDF format in the output directory."""
        try:
            # Create output directory if it doesn't exist
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Create output PDF path
            output_pdf = output_dir / f"{input_path.stem}.pdf"
            
            # Handle .eml files specifically
            if input_path.suffix.lower() == '.eml':
                try:
                    with open(input_path, 'rb') as f:
                        msg = BytesParser(policy=policy.default).parse(f)
                    
                    # Create PDF
                    pdf = FPDF()
                    pdf.add_page()
                    
                    # Set font and size
                    pdf.set_font('Helvetica', '', 10)
                    
                    # Set margins
                    pdf.set_margins(15, 15, 15)
                    
                    # Get content and split into lines
                    content = self.extract_email_content(msg)
                    lines = content.split('\n')
                    
                    # Add content to PDF
                    for line in lines:
                        # Handle long lines by wrapping
                        if len(line) > 100:
                            words = line.split()
                            current_line = ""
                            for word in words:
                                if len(current_line) + len(word) + 1 <= 100:
                                    current_line += " " + word if current_line else word
                                else:
                                    pdf.multi_cell(0, 5, current_line)
                                    current_line = word
                            if current_line:
                                pdf.multi_cell(0, 5, current_line)
                        else:
                            pdf.multi_cell(0, 5, line)
                    
                    # Save PDF
                    pdf.output(str(output_pdf))
                    return output_pdf
                    
                except Exception as e:
                    self.logger.error(f"Error converting .eml file {input_path} to PDF: {str(e)}")
                    return None

            # Handle Excel files and CSV files with improved formatting
            elif input_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm', '.xlsb', '.csv']:
                try:
                    self.logger.info(f"Converting file: {input_path}")
                    
                    # Get data based on file type
                    if input_path.suffix.lower() == '.csv':
                        # Read CSV file
                        data = []
                        max_cols = 0
                        with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
                            reader = csv.reader(f)
                            # First pass: find the maximum number of columns
                            for row in reader:
                                max_cols = max(max_cols, len(row))
                                
                        # Second pass: pad rows to match max columns
                        with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
                            reader = csv.reader(f)
                            for row in reader:
                                # Pad row with empty strings if needed
                                padded_row = row + [''] * (max_cols - len(row))
                                data.append(padded_row)
                                
                        if not data or max_cols == 0:
                            self.logger.warning(f"No data found in CSV file: {input_path}")
                            return None
                            
                        sheets = [{'title': input_path.stem, 'data': data}]
                    else:  # Excel files
                        if input_path.suffix.lower() in ['.xlsx', '.xlsm', '.xlsb']:
                            wb = load_workbook(input_path, data_only=True)
                            sheets = []
                            for sheet in wb.worksheets:
                                sheet_data = []
                                max_row = sheet.max_row
                                max_col = sheet.max_column
                                
                                # Find last non-empty row and column
                                last_row = 0
                                last_col = 0
                                for row in range(1, max_row + 1):
                                    for col in range(1, max_col + 1):
                                        cell = sheet.cell(row=row, column=col)
                                        if cell.value is not None and str(cell.value).strip():
                                            last_row = max(last_row, row)
                                            last_col = max(last_col, col)
                                
                                # Only process if we found non-empty cells
                                if last_row > 0 and last_col > 0:
                                    for row in range(1, last_row + 1):
                                        row_data = []
                                        for col in range(1, last_col + 1):
                                            cell = sheet.cell(row=row, column=col)
                                            value = cell.value if cell.value is not None else ''
                                            row_data.append(str(value))
                                        sheet_data.append(row_data)
                                    sheets.append({'title': sheet.title, 'data': sheet_data})
                        else:  # .xls files
                            wb = xlrd.open_workbook(input_path)
                            sheets = []
                            for sheet_idx in range(wb.nsheets):
                                sheet = wb.sheet_by_index(sheet_idx)
                                sheet_data = []
                                # Find last non-empty row and column
                                last_row = 0
                                last_col = 0
                                for row in range(sheet.nrows):
                                    for col in range(sheet.ncols):
                                        value = sheet.cell_value(row, col)
                                        if value is not None and str(value).strip():
                                            last_row = max(last_row, row)
                                            last_col = max(last_col, col)
                                
                                # Only process if we found non-empty cells
                                if last_row > 0 and last_col > 0:
                                    sheet_data = [[str(sheet.cell_value(r, c)) for c in range(last_col + 1)] 
                                                for r in range(last_row + 1)]
                                    sheets.append({'title': sheet.name, 'data': sheet_data})
                    
                    # Create PDF with landscape orientation and adjusted margins
                    doc = SimpleDocTemplate(
                        str(output_pdf),
                        pagesize=landscape(A4),  # Use A4 for more space
                        leftMargin=0.25*inch,
                        rightMargin=0.25*inch,
                        topMargin=0.25*inch,
                        bottomMargin=0.25*inch
                    )
                    
                    elements = []
                    styles = getSampleStyleSheet()
                    
                    # Create custom styles
                    title_style = ParagraphStyle(
                        'CustomTitle',
                        parent=styles['Heading1'],
                        fontSize=12,  # Slightly smaller title
                        spaceAfter=10,
                        alignment=1,
                        textColor=colors.black
                    )
                    
                    header_style = ParagraphStyle(
                        'HeaderStyle',
                        parent=styles['Normal'],
                        fontSize=9,  # Slightly smaller header
                        leading=10,
                        alignment=1,
                        textColor=colors.white,
                        fontName='Helvetica-Bold'
                    )
                    
                    cell_style = ParagraphStyle(
                        'CellStyle',
                        parent=styles['Normal'],
                        fontSize=8,  # Slightly smaller cell text
                        leading=9,
                        alignment=0,
                        textColor=colors.black,
                        fontName='Helvetica'
                    )
                    
                    # Process each sheet/data set
                    for sheet in sheets:
                        if not sheet['data']:
                            continue
                            
                        self.logger.info(f"Processing sheet: {sheet['title']}")
                        data = sheet['data']
                        
                        if data:
                            # Add sheet title (except for single CSV files)
                            if input_path.suffix.lower() != '.csv':
                                elements.append(Paragraph(sheet['title'], title_style))
                                elements.append(Spacer(1, 5))
                            
                            # Calculate optimal column widths
                            available_width = doc.width * 0.95
                            col_widths = []
                            min_col_width = 0.4 * inch  # Reduced minimum width
                            max_col_width = 1.5 * inch  # Reduced maximum width
                            
                            # First pass: calculate ideal widths
                            for col in range(len(data[0])):
                                max_width = 0
                                for row in data:
                                    if col < len(row):
                                        cell_text = str(row[col])
                                        # Calculate width based on content
                                        words = cell_text.split()
                                        if words:
                                            max_word_len = max(len(word) for word in words)
                                            total_len = len(cell_text)
                                            width = max(max_word_len * 0.12, (total_len * 0.1) / 2)
                                        else:
                                            width = len(cell_text) * 0.1
                                        max_width = max(max_width, width)
                                col_width = max(min_col_width, min(max_col_width, max_width * inch))
                                col_widths.append(col_width)
                            
                            # Adjust if total width exceeds available width
                            total_width = sum(col_widths)
                            if total_width > available_width:
                                scale_factor = available_width / total_width
                                col_widths = [max(min_col_width, w * scale_factor) for w in col_widths]
                            
                            # Split data into chunks that fit on a page
                            max_rows_per_page = 30  # Adjust based on content
                            data_chunks = [data[i:i + max_rows_per_page] for i in range(0, len(data), max_rows_per_page)]
                            
                            for chunk_idx, chunk in enumerate(data_chunks):
                                # Process data with improved text wrapping
                                wrapped_data = []
                                for row_idx, row in enumerate(chunk):
                                    wrapped_row = []
                                    for col_idx, cell in enumerate(row):
                                        cell_text = str(cell).strip()
                                        
                                        # Create appropriate style based on row
                                        style = header_style if (chunk_idx == 0 and row_idx == 0) else cell_style
                                        
                                        # Handle special number formatting
                                        if row_idx > 0 or chunk_idx > 0:  # Don't format headers
                                            try:
                                                if cell_text.replace(',', '').replace('.', '').replace('-', '').isdigit():
                                                    num = float(cell_text.replace(',', ''))
                                                    cell_text = "{:,.2f}".format(num)
                                            except:
                                                pass
                                        
                                        # Calculate available width for wrapping
                                        avail_width = col_widths[col_idx] - 6  # Reduced padding
                                        
                                        # Smart word wrapping with max lines
                                        words = cell_text.split()
                                        lines = []
                                        current_line = []
                                        current_width = 0
                                        max_lines = 3  # Limit number of lines per cell
                                        
                                        for word in words:
                                            word_width = len(word) * style.fontSize * 0.6
                                            if current_width + word_width <= avail_width and len(lines) < max_lines:
                                                current_line.append(word)
                                                current_width += word_width + style.fontSize * 0.3
                                            else:
                                                if current_line:
                                                    lines.append(' '.join(current_line))
                                                if len(lines) >= max_lines:
                                                    break
                                                current_line = [word]
                                                current_width = word_width
                                        
                                        if current_line and len(lines) < max_lines:
                                            lines.append(' '.join(current_line))
                                        
                                        # Join lines with HTML line breaks
                                        final_text = '<br/>'.join(lines)
                                        
                                        # Create paragraph with appropriate style
                                        p = Paragraph(final_text, style)
                                        wrapped_row.append(p)
                                    
                                    wrapped_data.append(wrapped_row)
                                
                                # Create table with calculated widths
                                table = Table(wrapped_data, colWidths=col_widths, repeatRows=1)
                                
                                # Add style with improved formatting
                                table.setStyle(TableStyle([
                                    # Header style
                                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                                    # Data style
                                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                                    # Alternate row colors
                                    *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2F2F2')) 
                                      for i in range(2, len(wrapped_data), 2)],
                                    # Grid style
                                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                                    ('BOX', (0, 0), (-1, -1), 1, colors.black),
                                    # Word wrap and alignment
                                    ('WORDWRAP', (0, 0), (-1, -1), True),
                                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                                ]))
                                
                                elements.append(table)
                                if chunk_idx < len(data_chunks) - 1:
                                    elements.append(PageBreak())
                    
                    # Only build PDF if there's content
                    if elements:
                        self.logger.info(f"Building PDF for {input_path}")
                        doc.build(elements)
                        self.logger.info(f"Successfully created PDF: {output_pdf}")
                        return output_pdf
                    else:
                        self.logger.warning(f"No data found in {input_path}")
                        return None
                    
                except Exception as e:
                    self.logger.error(f"Error converting file {input_path} to PDF: {str(e)}")
                    self.logger.error(f"Error details: {type(e).__name__}: {str(e)}")
                    return None

            # For other file types, try to convert using unoconv
            else:
                try:
                    if platform.system() == 'Darwin':  # macOS
                        cmd = [
                            'unoconv',
                            '-f', 'pdf',
                            '-o', str(output_pdf),
                            str(input_path)
                        ]
                    else:
                        cmd = ['unoconv', '-f', 'pdf', '-o', str(output_pdf), str(input_path)]
                    
                    # Run the conversion command
                    result = subprocess.run(cmd, capture_output=True, text=True)
                    
                    if result.returncode != 0:
                        self.logger.error(f"Error converting {input_path} to PDF: {result.stderr}")
                        return None
                    
                    return output_pdf
                    
                except Exception as e:
                    self.logger.error(f"Error converting {input_path} to PDF: {str(e)}")
                    return None
            
        except Exception as e:
            self.logger.error(f"Error converting {input_path} to PDF: {str(e)}")
            return None

    def add_bates_stamp(self, input_pdf: Path, output_pdf: Path, bates_number: str) -> bool:
        """Add Bates number stamp to each page of the PDF."""
        try:
            # Try to unlock the PDF if it's locked
            try:
                # Create a temporary copy of the input PDF to handle secured PDFs
                temp_pdf = input_pdf.parent / f"temp_{input_pdf.name}"
                try:
                    # Try to create an unsecured copy of the PDF
                    reader = PdfReader(input_pdf)
                    writer = PdfWriter()
                    
                    # Copy all pages to remove security
                    for page in reader.pages:
                        writer.add_page(page)
                    
                    # Write unsecured temporary PDF
                    with open(temp_pdf, 'wb') as temp_file:
                        writer.write(temp_file)
                    
                    # Now use the unsecured PDF for stamping
                    reader = PdfReader(temp_pdf)
                except Exception as e:
                    self.logger.warning(f"Could not unlock PDF {input_pdf}, trying to print to PDF: {str(e)}")
                    # If unlocking fails, try to print to PDF
                    if platform.system() == 'Darwin':  # macOS
                        cmd = [
                            'cupsfilter',
                            '-o', 'fit-to-page',
                            '-o', 'media=letter',
                            str(input_pdf)
                        ]
                        result = subprocess.run(cmd, capture_output=True, text=True)
                        if result.returncode == 0:
                            temp_pdf = input_pdf.parent / f"printed_{input_pdf.name}"
                            with open(temp_pdf, 'wb') as f:
                                f.write(result.stdout.encode())
                            reader = PdfReader(temp_pdf)
                        else:
                            raise Exception("Failed to print PDF")
                    else:
                        raise Exception("PDF unlocking and printing not supported on this platform")
            except Exception as e:
                self.logger.error(f"Error handling locked PDF {input_pdf}: {str(e)}")
                return False
            
            writer = PdfWriter()
            
            # Track if we successfully processed any pages
            pages_processed = 0
            total_pages = len(reader.pages)
            
            # Store the starting Bates number for this file
            current_bates = self.current_number
            
            # Process each page
            for page_num, page in enumerate(reader.pages, 1):
                try:
                    # Get page dimensions
                    width = float(page.mediabox.width)
                    height = float(page.mediabox.height)
                    
                    # Create a new PDF with the Bates number
                    packet = io.BytesIO()
                    can = canvas.Canvas(packet, pagesize=(width, height))
                    can.setFont("Helvetica-Bold", 10)
                    
                    # Set color based on settings
                    if self.stamp_color == "black":
                        fill_color = black
                    elif self.stamp_color == "red":
                        fill_color = colors.red
                    elif self.stamp_color == "blue":
                        fill_color = colors.blue
                    elif self.stamp_color == "green":
                        fill_color = colors.green
                    elif self.stamp_color == "gray":
                        fill_color = colors.gray
                    
                    # Apply opacity
                    can.setFillColor(fill_color, alpha=self.stamp_opacity)
                    
                    # Generate Bates number for this page
                    if page_num == 1:
                        # For first page, use the provided Bates number
                        page_bates = bates_number
                    else:
                        # For subsequent pages, increment from the starting number
                        page_bates = f"{self.prefix}{str(self.start + page_num - 1).zfill(self.zero_pad_length)}"
                    
                    # Calculate text dimensions
                    text_width = can.stringWidth(page_bates, "Helvetica-Bold", 10)
                    text_height = 10
                    
                    # Set default margins
                    margin = 10
                    
                    # Calculate position based on grid selection
                    if "top" in self.stamp_position:
                        y = height - margin - text_height
                    elif "middle" in self.stamp_position:
                        y = height / 2
                    else:  # bottom
                        y = margin + text_height
                    
                    if "left" in self.stamp_position:
                        x = margin + text_width
                    elif "center" in self.stamp_position:
                        x = width / 2
                    else:  # right
                        x = width - margin
                    
                    # Apply offsets
                    x += self.stamp_x_offset
                    y += self.stamp_y_offset
                    
                    # Draw box if width > 0
                    if self.stamp_box_width > 0:
                        # Calculate box dimensions based on position
                        box_padding = 2
                        box_width = text_width + (box_padding * 2)
                        box_height = text_height + (box_padding * 2)
                        
                        if "left" in self.stamp_position:
                            box_x = x - text_width - box_padding
                        elif "center" in self.stamp_position:
                            box_x = x - (box_width / 2)
                        else:  # right
                            box_x = x - text_width - box_padding
                        
                        if "top" in self.stamp_position:
                            box_y = y - text_height - box_padding
                        elif "middle" in self.stamp_position:
                            box_y = y - (box_height / 2)
                        else:  # bottom
                            box_y = y - box_padding
                        
                        # Draw rectangle around text
                        can.setStrokeColor(fill_color, alpha=self.stamp_opacity)
                        can.setLineWidth(self.stamp_box_width)
                        can.rect(box_x, box_y, box_width, box_height)
                        
                        # Draw text based on position
                        if "left" in self.stamp_position:
                            can.drawString(box_x + box_padding, box_y + box_padding, page_bates)
                        elif "center" in self.stamp_position:
                            can.drawCentredString(x, box_y + box_padding, page_bates)
                        else:  # right
                            can.drawRightString(box_x + box_width - box_padding, box_y + box_padding, page_bates)
                    else:
                        # Draw text without box
                        if "left" in self.stamp_position:
                            can.drawString(x - text_width, y, page_bates)
                        elif "center" in self.stamp_position:
                            can.drawCentredString(x, y, page_bates)
                        else:  # right
                            can.drawRightString(x, y, page_bates)
                    
                    can.save()
                    packet.seek(0)
                    
                    # Create a new PDF with the Bates number
                    new_pdf = PdfReader(packet)
                    
                    # Create a new page object
                    new_page = page
                    
                    try:
                        # Try to merge the Bates number with the page
                        new_page.merge_page(new_pdf.pages[0])
                    except Exception as e:
                        # If merging fails, try alternative method
                        self.logger.warning(f"Warning: Primary merge failed for page {page_num}, trying alternative method: {str(e)}")
                        try:
                            # Create a new blank page
                            packet = io.BytesIO()
                            can = canvas.Canvas(packet, pagesize=(width, height))
                            
                            # Copy the original page content
                            can.doForm(page['/Resources']['/XObject'].keys()[0])
                            
                            # Add the Bates number with box
                            can.setFont("Helvetica-Bold", 10)
                            can.setFillColor(can._fillColor, alpha=self.stamp_opacity)
                            
                            # Draw box if width > 0
                            if self.stamp_box_width > 0:
                                box_padding = 2
                                box_width = text_width + (box_padding * 2)
                                box_height = text_height + (box_padding * 2)
                                
                                if "left" in self.stamp_position:
                                    box_x = x - text_width - box_padding
                                elif "center" in self.stamp_position:
                                    box_x = x - (box_width / 2)
                                else:  # right
                                    box_x = x - text_width - box_padding
                                
                                if "top" in self.stamp_position:
                                    box_y = y - text_height - box_padding
                                elif "middle" in self.stamp_position:
                                    box_y = y - (box_height / 2)
                                else:  # bottom
                                    box_y = y - box_padding
                                
                                can.setStrokeColor(can._fillColor, alpha=self.stamp_opacity)
                                can.setLineWidth(self.stamp_box_width)
                                can.rect(box_x, box_y, box_width, box_height)
                                
                                if "left" in self.stamp_position:
                                    can.drawString(box_x + box_padding, box_y + box_padding, page_bates)
                                elif "center" in self.stamp_position:
                                    can.drawCentredString(x, box_y + box_padding, page_bates)
                                else:  # right
                                    can.drawRightString(box_x + box_width - box_padding, box_y + box_padding, page_bates)
                            else:
                                if "left" in self.stamp_position:
                                    can.drawString(x - text_width, y, page_bates)
                                elif "center" in self.stamp_position:
                                    can.drawCentredString(x, y, page_bates)
                                else:  # right
                                    can.drawRightString(x, y, page_bates)
                            
                            can.save()
                            packet.seek(0)
                            
                            # Create new page with combined content
                            new_page = PdfReader(packet).pages[0]
                        except Exception as e2:
                            self.logger.warning(f"Warning: Alternative method also failed for page {page_num}: {str(e2)}")
                            # If both methods fail, just add the original page
                            new_page = page
                    
                    writer.add_page(new_page)
                    pages_processed += 1
                    
                except Exception as e:
                    # Log a warning for page processing errors
                    self.logger.warning(f"Warning processing page {page_num} in {input_pdf}: {str(e)}")
                    # Add the page without stamping if there's an error
                    writer.add_page(page)
                    pages_processed += 1
            
            # Only proceed if we processed at least one page
            if pages_processed > 0:
                # Create the output directory if it doesn't exist
                output_pdf.parent.mkdir(parents=True, exist_ok=True)
                
                # Write the stamped PDF to a temporary location
                temp_stamped = output_pdf.parent / f"temp_stamped_{output_pdf.name}"
                with open(temp_stamped, 'wb') as output_file:
                    writer.write(output_file)
                
                # Move the stamped file to its final location
                shutil.move(temp_stamped, output_pdf)
                
                # Log success with page count
                self.logger.info(f"Successfully processed {pages_processed} of {total_pages} pages in {input_pdf}")
                
                # Update the current number for the next file
                self.current_number = self.start + total_pages
                return True
            else:
                self.logger.error(f"Failed to process any pages in {input_pdf}")
                return False
            
        finally:
            # Clean up temporary files
            if 'temp_pdf' in locals() and temp_pdf.exists():
                temp_pdf.unlink()
            if 'temp_stamped' in locals() and temp_stamped.exists():
                temp_stamped.unlink()

    def should_ignore_file(self, file_path: Path) -> bool:
        """Check if a file should be ignored based on its name or extension."""
        # Check file name patterns
        for pattern in self.ignored_files:
            if pattern in file_path.name:
                self.logger.debug(f"Ignoring file matching pattern {pattern}: {file_path}")
                return True
        
        # Check file extension
        if file_path.suffix.lower() in self.ignored_files:
            self.logger.debug(f"Ignoring file with extension {file_path.suffix}: {file_path}")
            return True
            
        return False

    def get_pdf_page_count(self, pdf_path: Path) -> int:
        """Get the number of pages in a PDF file."""
        try:
            with open(pdf_path, 'rb') as file:
                reader = PdfReader(file)
                return len(reader.pages)
        except Exception as e:
            self.logger.error(f"Error getting page count for {pdf_path}: {str(e)}")
            return 0

    def create_combined_pdf(self):
        """Create a combined PDF of all processed files."""
        try:
            merger = PdfMerger()
            combined_path = self.output_dir / f"{self.prefix}_combined.pdf"
            
            # Get all PDF files and sort them by Bates number
            pdf_files = []
            for pdf_file in self.output_dir.glob("**/*.pdf"):
                if "combined" not in pdf_file.name.lower() and "issues" not in pdf_file.name.lower():
                    bates_match = re.search(rf"{self.prefix}\d+", pdf_file.name)
                    if bates_match:
                        bates_number = bates_match.group()
                        pdf_files.append((bates_number, pdf_file))
            
            # Sort files by Bates number
            pdf_files.sort(key=lambda x: int(re.search(r'\d+', x[0]).group()))
            
            # Track failed files
            failed_files = []
            
            # Merge PDFs with error handling
            for bates_number, pdf_file in pdf_files:
                try:
                    # Verify PDF is not corrupted before adding
                    with open(pdf_file, 'rb') as f:
                        reader = PdfReader(f)
                        if len(reader.pages) > 0:  # Basic check for valid PDF
                            merger.append(str(pdf_file))
                        else:
                            self.logger.warning(f"Skipping empty PDF: {pdf_file}")
                            failed_files.append(pdf_file)
                except Exception as e:
                    self.logger.warning(f"Error processing {pdf_file}, skipping: {str(e)}")
                    failed_files.append(pdf_file)
            
            # Only write the combined PDF if we successfully added any files
            if len(merger.pages) > 0:
                merger.write(str(combined_path))
                self.logger.info(f"Combined PDF created: {combined_path}")
                
                # Log any failed files
                if failed_files:
                    self.logger.warning(f"Failed to include {len(failed_files)} files in combined PDF")
                    for failed_file in failed_files:
                        self.logger.warning(f"Failed file: {failed_file}")
            else:
                self.logger.warning("No valid PDFs found to combine")
            
            merger.close()
            
        except Exception as e:
            self.logger.error(f"Error creating combined PDF: {str(e)}")
            # Don't raise the error, just log it and continue
            self.logger.info("Continuing without combined PDF")

    def copy_source_file(self, source_path: Path, bates_number: str) -> bool:
        """Copy source file to output directory with Bates number prefix."""
        try:
            # Create the relative path structure in the output directory
            rel_path = source_path.relative_to(self.input_dir)
            target_dir = self.output_dir / rel_path.parent
            target_dir.mkdir(parents=True, exist_ok=True)
            
            # Create new filename with Bates number
            new_filename = f"{bates_number}_{source_path.name}"
            target_path = target_dir / new_filename
            
            # Copy the file
            shutil.copy2(source_path, target_path)
            self.logger.info(f"Copied source file to: {target_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error copying source file {source_path}: {str(e)}")
            return False

    def extract_email_content(self, msg):
        """Extract email content, handling both plain text and HTML."""
        content = []
        
        # Get email metadata
        subject = msg.get('subject', 'No Subject')
        from_addr = msg.get('from', 'Unknown Sender')
        to_addr = msg.get('to', 'Unknown Recipient')
        date = msg.get('date', 'Unknown Date')
        
        content.append(f"From: {from_addr}")
        content.append(f"To: {to_addr}")
        content.append(f"Subject: {subject}")
        content.append(f"Date: {date}")
        content.append("\n" + "="*50 + "\n")
        
        # Process email body
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    try:
                        text = part.get_content()
                        if isinstance(text, bytes):
                            text = text.decode()
                        content.append(text)
                    except Exception as e:
                        content.append(f"Error decoding text part: {str(e)}")
                elif part.get_content_type() == "text/html":
                    try:
                        html_content = part.get_content()
                        if isinstance(html_content, bytes):
                            html_content = html_content.decode()
                        # Basic HTML to text conversion
                        html_content = re.sub('<[^<]+?>', '', html_content)
                        html_content = html.unescape(html_content)
                        content.append(html_content)
                    except Exception as e:
                        content.append(f"Error decoding HTML part: {str(e)}")
        else:
            try:
                text = msg.get_content()
                if isinstance(text, bytes):
                    text = text.decode()
                content.append(text)
            except Exception as e:
                content.append(f"Error decoding content: {str(e)}")
        
        return "\n".join(content)

    def process_files(self):
        """Process all PDF files in the input directory."""
        try:
            # Create output directory if it doesn't exist
            self.output_dir.mkdir(parents=True, exist_ok=True)
            
            # Track processed files to avoid duplicates
            processed_files = set()
            
            # Process each file in the input directory
            for input_file in sorted(self.input_dir.glob("**/*")):
                if not input_file.is_file() or self.should_ignore_file(input_file):
                    continue
                
                # Skip if we've already processed this file
                if str(input_file) in processed_files:
                    continue
                
                try:
                    # Generate Bates number for this file
                    bates_number = f"{self.prefix}{str(self.current_number).zfill(self.zero_pad_length)}"
                    
                    # Create the relative path structure in the output directory
                    rel_path = input_file.relative_to(self.input_dir)
                    target_dir = self.output_dir / rel_path.parent
                    target_dir.mkdir(parents=True, exist_ok=True)
                    
                    # If it's a PDF, copy it and add Bates stamps
                    if input_file.suffix.lower() == '.pdf':
                        # Copy the original PDF with Bates number prefix
                        new_filename = f"{bates_number}_{input_file.name}"
                        target_path = target_dir / new_filename
                        shutil.copy2(input_file, target_path)
                        self.logger.info(f"Copied source file to: {target_path}")

                        # Add Bates stamps to the copied PDF
                        if not self.add_bates_stamp(input_file, target_path, bates_number):
                            self.move_to_issues(target_path, "Failed to apply Bates stamp")
                            continue
                    else:
                        # For non-PDF files:
                        if input_file.suffix.lower() == '.eml':
                            # For EML files:
                            # 1. Copy the original EML file with Bates number prefix
                            original_new_filename = f"{bates_number}_{input_file.name}"
                            original_target_path = target_dir / original_new_filename
                            shutil.copy2(input_file, original_target_path)
                            self.logger.info(f"Copied original EML file to: {original_target_path}")
                            
                            # 2. Create PDF version with the same Bates number
                            pdf_filename = f"{bates_number}_{input_file.stem}.pdf"
                            pdf_path = target_dir / pdf_filename
                            
                            # Convert to PDF
                            converted_pdf = self.convert_to_pdf(input_file, target_dir)
                            if not converted_pdf:
                                self.logger.error(f"Failed to convert {input_file} to PDF")
                                continue
                                
                            # Move the converted PDF to the final location with Bates number
                            shutil.move(str(converted_pdf), str(pdf_path))
                            
                            # Add Bates stamps to the PDF
                            if not self.add_bates_stamp(pdf_path, pdf_path, bates_number):
                                self.move_to_issues(pdf_path, "Failed to apply Bates stamp")
                                continue
                        else:
                            # For other non-PDF files:
                            # 1. Copy the original file with Bates number prefix
                            original_new_filename = f"{bates_number}_{input_file.name}"
                            original_target_path = target_dir / original_new_filename
                            shutil.copy2(input_file, original_target_path)
                            self.logger.info(f"Copied original file to: {original_target_path}")
                            
                            # 2. Create PDF version with the same Bates number
                            pdf_filename = f"{bates_number}_{input_file.stem}.pdf"
                            pdf_path = target_dir / pdf_filename
                            
                            # Convert to PDF
                            converted_pdf = self.convert_to_pdf(input_file, target_dir)
                            if not converted_pdf:
                                self.logger.error(f"Failed to convert {input_file} to PDF")
                                continue
                                
                            # Move the converted PDF to the final location with Bates number
                            shutil.move(str(converted_pdf), str(pdf_path))
                            
                            # Add Bates stamps to the PDF
                            if not self.add_bates_stamp(pdf_path, pdf_path, bates_number):
                                self.move_to_issues(pdf_path, "Failed to apply Bates stamp")
                                continue

                    # Mark file as processed
                    processed_files.add(str(input_file))
                    
                except Exception as e:
                    self.logger.error(f"Error processing {input_file}: {str(e)}")
                    if 'target_path' in locals():
                        self.move_to_issues(target_path, str(e))
                    elif 'pdf_path' in locals():
                        self.move_to_issues(pdf_path, str(e))

        except Exception as e:
            self.logger.error(f"Error during file processing: {str(e)}")
            raise

    def generate_excel(self):
        """Generate Excel report of processed files."""
        try:
            excel_path = self.output_dir / "bates_report.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Bates Report"

            # Add headers
            headers = ["Bates Number", "Original Filename", "Page Count", "Creation Date", "Processing Date"]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)

            # Add data
            row = 2
            for pdf_file in sorted(self.output_dir.glob("**/*.pdf")):
                if "combined" not in pdf_file.name.lower():
                    bates_match = re.search(rf"{self.prefix}\d+", pdf_file.name)
                    bates_number = bates_match.group() if bates_match else "N/A"
                    
                    sheet.cell(row=row, column=1, value=bates_number)
                    sheet.cell(row=row, column=2, value=pdf_file.name)
                    sheet.cell(row=row, column=3, value=self.get_pdf_page_count(pdf_file))
                    sheet.cell(row=row, column=4, value=datetime.fromtimestamp(pdf_file.stat().st_ctime).strftime('%Y-%m-%d'))
                    sheet.cell(row=row, column=5, value=datetime.now().strftime('%Y-%m-%d'))
                    row += 1

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            workbook.save(excel_path)
            self.logger.info(f"Excel report generated: {excel_path}")

        except Exception as e:
            self.logger.error(f"Error generating Excel report: {str(e)}")
            raise

    def run(self):
        """Run the Bates numbering process."""
        try:
            # Process files
            self.process_files()
            
            # Generate Excel report
            self.generate_excel()
            
            # Create combined PDF
            self.create_combined_pdf()
            
            self.logger.info(f"Processing complete. Output directory: {self.output_dir}")
            
        except Exception as e:
            self.logger.error(f"Error during processing: {str(e)}")
            raise

    def get_pdf_files(self) -> List[Tuple[Path, Path]]:
        """Get all PDF files and convertible files in the input directory and its subdirectories.
        Returns a list of tuples (file_path, pdf_path) where pdf_path is None for native PDFs."""
        files_to_process = []
        try:
            # Walk through the directory tree
            for root, _, files in os.walk(self.input_dir):
                for file in files:
                    file_path = Path(root) / file
                    
                    # Skip if file should be ignored
                    if self.should_ignore_file(file_path):
                        continue
                    
                    # If it's a PDF, add it directly
                    if file_path.suffix.lower() == '.pdf':
                        files_to_process.append((file_path, None))
                    # If it's a convertible file, add it for conversion
                    elif file_path.suffix.lower() in ['.docx', '.doc', '.xlsx', '.xls', '.png', '.jpg', '.jpeg', '.gif', '.bmp', '.eml']:
                        # Create the relative path structure in the output directory
                        rel_path = file_path.relative_to(self.input_dir)
                        output_dir = self.output_dir / rel_path.parent
                        files_to_process.append((file_path, output_dir))
            
            # Sort files by name for consistent processing
            files_to_process.sort(key=lambda x: str(x[0]))
            
            if not files_to_process:
                self.logger.warning("No files found to process")
            else:
                self.logger.info(f"Found {len(files_to_process)} files to process")
            
            return files_to_process
            
        except Exception as e:
            self.logger.error(f"Error getting files: {str(e)}")
            return []

    def is_ignored_file(self, file_path: Path) -> bool:
        """Check if a file should be ignored based on its name or extension."""
        # Check file name patterns
        for pattern in self.ignored_files:
            if pattern in file_path.name:
                self.logger.debug(f"Ignoring file matching pattern {pattern}: {file_path}")
                return True
        
        # Check file extension
        if file_path.suffix.lower() in self.ignored_files:
            self.logger.debug(f"Ignoring file with extension {file_path.suffix}: {file_path}")
            return True
            
        return False

    def extract_date_from_filename(self, filename: str) -> datetime:
        """Extract date from filename if present."""
        try:
            # Common date patterns in filenames
            date_patterns = [
                r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
                r'\d{2}-\d{2}-\d{4}',  # MM-DD-YYYY
                r'\d{8}',              # YYYYMMDD
                r'\d{6}'               # YYMMDD
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, filename)
                if match:
                    date_str = match.group()
                    try:
                        if len(date_str) == 8:  # YYYYMMDD
                            return datetime.strptime(date_str, '%Y%m%d')
                        elif len(date_str) == 6:  # YYMMDD
                            return datetime.strptime(date_str, '%y%m%d')
                        elif '-' in date_str:
                            if len(date_str.split('-')[0]) == 4:  # YYYY-MM-DD
                                return datetime.strptime(date_str, '%Y-%m-%d')
                            else:  # MM-DD-YYYY
                                return datetime.strptime(date_str, '%m-%d-%Y')
                    except ValueError:
                        continue
            
            return None
            
        except Exception as e:
            self.logger.warning(f"Error extracting date from filename {filename}: {str(e)}")
            return None

    def move_to_issues(self, file_path: Path, reason: str):
        """Move a file to the __FILES WITH ISSUES folder with a reason."""
        try:
            # Create issues drectory if it doesn't exist
            self.issues_dir.mkdir(parents=True, exist_ok=True)
            
            # Create a new filename with the reason
            new_name = f"{file_path.stem}_ISSUE_{reason.replace(' ', '_')}{file_path.suffix}"
            new_path = self.issues_dir / new_name
            
            # Move the file
            shutil.move(str(file_path), str(new_path))
            self.logger.info(f"Moved {file_path} to issues folder: {reason}")
            
        except Exception as e:
            self.logger.error(f"Error moving {file_path} to issues folder: {str(e)}")

class BatesGUI:
    def __init__(self):
        if HAS_TKINTERDND:
            self.root = TkinterDnD.Tk()
        else:
            self.root = tk.Tk()
        self.root.title("Bates Numbering / Stamping Utility")
        
        self.root.attributes('-topmost', True)
        
        # Make window resizable
        self.root.resizable(True, True)
        
        # Set minimum window size to prevent buttons from being hidden
        self.root.minsize(1200, 400)
        
        # Variables
        self.input_dir = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.input_file = tk.StringVar()
        self.file_output_dir = tk.StringVar()
        self.prefix = tk.StringVar(value="ABC")  # Default prefix
        self.digits = tk.IntVar(value=5)  # Default to 5 digits
        self.start_number = tk.IntVar(value=1)  # Default starting number
        
        # Stamp appearance settings
        self.stamp_x = tk.DoubleVar(value=0.97)  # Default 97% from left (3% from right)
        self.stamp_y = tk.DoubleVar(value=0.001)  # Default 0.1% from bottom
        self.stamp_color = tk.StringVar(value="black")  # Default color
        self.stamp_box_width = tk.DoubleVar(value=1.0)  # Default box width of 1
        self.stamp_position = tk.StringVar(value="bottom-right")  # Default position
        self.stamp_x_offset = tk.IntVar(value=0)  # Default X offset
        self.stamp_y_offset = tk.IntVar(value=0)  # Default Y offset
        self.stamp_opacity = tk.IntVar(value=100)  # Default opacity of 100%
        
        self.processing = False
        self.last_directory = str(Path.home())  # Initialize with home directory
        
        # Set default input directory
        default_dir = Path('~/Desktop')
        if default_dir.exists():
            self.input_dir.set(str(default_dir))
            # Set default output directory to BATES_ + input directory name
            self.output_dir.set(str(default_dir.parent / f"BATES_{default_dir.name}"))
        
        self.create_widgets()
        
        # Position window at the top of the screen after widgets are created
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = 0
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        # Setup drag and drop after widgets are created
        if HAS_TKINTERDND:
            self.setup_drag_drop()
            
    def create_widgets(self):
        # Main container with padding
        self.main_frame = ttk.Frame(self.root, padding="5")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create Notebook (tabs)
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Create Directory Tab
        self.dir_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.dir_tab, text="Directory")
        
        # Create File Tab
        self.file_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.file_tab, text="File")
        
        # Create Directory Tab Content
        self.create_directory_tab()
        
        # Create File Tab Content
        self.create_file_tab()
        
        # Create Common Settings Frame (at bottom)
        self.create_common_settings()
        
    def create_directory_tab(self):
        # Directory Tab Content
        dir_frame = ttk.Frame(self.dir_tab, padding="5")
        dir_frame.pack(fill=tk.BOTH, expand=True)
        
        # Source Directory Section
        source_frame = ttk.LabelFrame(dir_frame, text="Source Directory", padding="5")
        source_frame.pack(fill=tk.X, pady=5)
        
        # Source Directory Drop Zone
        source_drop_zone = ttk.Frame(source_frame)
        source_drop_zone.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Center container for source label
        source_center_frame = ttk.Frame(source_drop_zone)
        source_center_frame.pack(expand=True)
        
        self.source_label = ttk.Label(source_center_frame, text="Drop folder here or use browse button", 
                                    wraplength=700, justify=tk.CENTER)
        self.source_label.pack(expand=True)
        
        # Source Directory Browse Button
        source_browse_frame = ttk.Frame(source_frame)
        source_browse_frame.pack(fill=tk.X, pady=5)
        ttk.Button(source_browse_frame, text="Browse Source", 
                  command=self.browse_folder).pack(expand=True)
        
        # Destination Directory Section
        dest_frame = ttk.LabelFrame(dir_frame, text="Destination Directory", padding="5")
        dest_frame.pack(fill=tk.X, pady=5)
        
        # Destination Directory Drop Zone
        dest_drop_zone = ttk.Frame(dest_frame)
        dest_drop_zone.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Center container for destination label
        dest_center_frame = ttk.Frame(dest_drop_zone)
        dest_center_frame.pack(expand=True)
        
        self.dest_label = ttk.Label(dest_center_frame, text="Drop destination folder here or use browse button.  (Defaults to Source Directory)", 
                                  wraplength=700, justify=tk.CENTER)
        self.dest_label.pack(expand=True)
        
        # Destination Directory Browse Button
        dest_browse_frame = ttk.Frame(dest_frame)
        dest_browse_frame.pack(fill=tk.X, pady=5)
        ttk.Button(dest_browse_frame, text="Browse Destination", 
                  command=self.browse_output_folder).pack(expand=True)
        
        # Stamp Directory Button
        stamp_dir_frame = ttk.Frame(dir_frame)
        stamp_dir_frame.pack(fill=tk.X, pady=5)
        self.stamp_dir_button = ttk.Button(stamp_dir_frame, text="Stamp Directory", 
                                         command=self.start_processing)
        self.stamp_dir_button.pack(expand=True)
        
    def create_file_tab(self):
        # File Tab Content
        file_frame = ttk.Frame(self.file_tab, padding="5")
        file_frame.pack(fill=tk.BOTH, expand=True)
        
        # File Selection Section
        file_select_frame = ttk.LabelFrame(file_frame, text="Select File", padding="5")
        file_select_frame.pack(fill=tk.X, pady=5)
        
        # File Drop Zone
        file_drop_zone = ttk.Frame(file_select_frame)
        file_drop_zone.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Center container for file label
        file_center_frame = ttk.Frame(file_drop_zone)
        file_center_frame.pack(expand=True)
        
        self.file_label = ttk.Label(file_center_frame, text="Drop file here or use browse button", 
                                  wraplength=700, justify=tk.CENTER)
        self.file_label.pack(expand=True)
        
        # File Browse Button
        file_browse_frame = ttk.Frame(file_select_frame)
        file_browse_frame.pack(fill=tk.X, pady=5)
        ttk.Button(file_browse_frame, text="Browse File", 
                  command=self.browse_file).pack(expand=True)
        
        # File Destination Section
        file_dest_frame = ttk.LabelFrame(file_frame, text="Destination", padding="5")
        file_dest_frame.pack(fill=tk.X, pady=5)
        
        # File Destination Drop Zone
        file_dest_drop_zone = ttk.Frame(file_dest_frame)
        file_dest_drop_zone.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Center container for file destination label
        file_dest_center_frame = ttk.Frame(file_dest_drop_zone)
        file_dest_center_frame.pack(expand=True)
        
        self.file_dest_label = ttk.Label(file_dest_center_frame, text="Drop destination folder here or use browse button.  (Defaults to Source Directory)", 
                                       wraplength=700, justify=tk.CENTER)
        self.file_dest_label.pack(expand=True)
        
        # File Destination Browse Button
        file_dest_browse_frame = ttk.Frame(file_dest_frame)
        file_dest_browse_frame.pack(fill=tk.X, pady=5)
        ttk.Button(file_dest_browse_frame, text="Browse Destination", 
                  command=self.browse_file_output_folder).pack(expand=True)
        
        # Stamp File Button
        stamp_file_frame = ttk.Frame(file_frame)
        stamp_file_frame.pack(fill=tk.X, pady=5)
        self.stamp_file_button = ttk.Button(stamp_file_frame, text="Stamp File", 
                                          command=self.stamp_single_file)
        self.stamp_file_button.pack(expand=True)
        
    def create_common_settings(self):
        # Common Settings Frame
        settings_frame = ttk.LabelFrame(self.main_frame, text="Settings", padding="5")
        settings_frame.pack(fill=tk.X, padx=20)
        
        # Create a container for the three columns
        columns_frame = ttk.Frame(settings_frame)
        columns_frame.pack(expand=True, pady=5)
        
        # Left Column - Base Settings
        base_frame = ttk.LabelFrame(columns_frame, text="Base Settings", padding="5")
        base_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        # Prefix
        prefix_frame = ttk.Frame(base_frame)
        prefix_frame.pack(fill=tk.X, pady=2)
        ttk.Label(prefix_frame, text="Prefix:", width=8).pack(side=tk.LEFT)
        ttk.Entry(prefix_frame, textvariable=self.prefix, width=10).pack(side=tk.LEFT)
        
        # Digits
        digits_frame = ttk.Frame(base_frame)
        digits_frame.pack(fill=tk.X, pady=2)
        ttk.Label(digits_frame, text="# of Digits:", width=8).pack(side=tk.LEFT)
        self.digits_spinbox = ttk.Spinbox(digits_frame, from_=1, to=10, 
                                        textvariable=self.digits,
                                        width=5, format="%0.0f")
        self.digits_spinbox.pack(side=tk.LEFT)
        
        # Start Number
        start_frame = ttk.Frame(base_frame)
        start_frame.pack(fill=tk.X, pady=2)
        ttk.Label(start_frame, text="Starting #:", width=8).pack(side=tk.LEFT)
        self.start_spinbox = ttk.Spinbox(start_frame, from_=1, to=999999, 
                                       textvariable=self.start_number,
                                       width=8, format="%0.0f")
        self.start_spinbox.pack(side=tk.LEFT)
        
        # Box Width (moved to Base Settings)
        box_frame = ttk.Frame(base_frame)
        box_frame.pack(fill=tk.X, pady=2)
        ttk.Label(box_frame, text="Box Width:", width=8).pack(side=tk.LEFT)
        self.box_spinbox = ttk.Spinbox(box_frame, from_=0, to=5, 
                                     textvariable=self.stamp_box_width,
                                     width=5, format="%0.1f", increment=0.1)
        self.box_spinbox.pack(side=tk.LEFT)
        
        # Center Column - Stamp Appearance
        appearance_frame = ttk.LabelFrame(columns_frame, text="Stamp Appearance", padding="5")
        appearance_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        # Color
        color_frame = ttk.Frame(appearance_frame)
        color_frame.pack(fill=tk.X, pady=2)
        ttk.Label(color_frame, text="Color:", width=8).pack(side=tk.LEFT)
        color_combo = ttk.Combobox(color_frame, textvariable=self.stamp_color, 
                                 values=["black", "red", "blue", "green", "gray"],
                                 width=8, state="readonly")
        color_combo.pack(side=tk.LEFT)
        
        # Opacity (new control)
        opacity_frame = ttk.Frame(appearance_frame)
        opacity_frame.pack(fill=tk.X, pady=2)
        ttk.Label(opacity_frame, text="Opacity:", width=8).pack(side=tk.LEFT)
        self.opacity_spinbox = ttk.Spinbox(opacity_frame, from_=0, to=100, 
                                         textvariable=self.stamp_opacity,
                                         width=5, format="%0.0f")
        self.opacity_spinbox.pack(side=tk.LEFT)
        
        # X Offset
        x_offset_frame = ttk.Frame(appearance_frame)
        x_offset_frame.pack(fill=tk.X, pady=2)
        ttk.Label(x_offset_frame, text="X Offset:", width=8).pack(side=tk.LEFT)
        self.x_offset_spinbox = ttk.Spinbox(x_offset_frame, from_=-100, to=100, 
                                          textvariable=self.stamp_x_offset,
                                          width=5, format="%0.0f")
        self.x_offset_spinbox.pack(side=tk.LEFT)
        
        # Y Offset
        y_offset_frame = ttk.Frame(appearance_frame)
        y_offset_frame.pack(fill=tk.X, pady=2)
        ttk.Label(y_offset_frame, text="Y Offset:", width=8).pack(side=tk.LEFT)
        self.y_offset_spinbox = ttk.Spinbox(y_offset_frame, from_=-100, to=100, 
                                          textvariable=self.stamp_y_offset,
                                          width=5, format="%0.0f")
        self.y_offset_spinbox.pack(side=tk.LEFT)
        
        # Right Column - Stamp Position
        position_frame = ttk.LabelFrame(columns_frame, text="Stamp Position", padding="5")
        position_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        # Create a 3x3 grid of buttons for position selection
        positions = {
            "top-left": (0, 0), "top-center": (0, 1), "top-right": (0, 2),
            "middle-left": (1, 0), "middle-center": (1, 1), "middle-right": (1, 2),
            "bottom-left": (2, 0), "bottom-center": (2, 1), "bottom-right": (2, 2)
        }
        
        # Create a frame for the grid
        grid_frame = ttk.Frame(position_frame)
        grid_frame.pack(expand=True)
        
        # Create and place the buttons
        self.position_buttons = {}
        for pos, (row, col) in positions.items():
            btn = ttk.Radiobutton(grid_frame, text="", value=pos, 
                                variable=self.stamp_position)
            btn.grid(row=row, column=col, padx=2, pady=2)
            self.position_buttons[pos] = btn
        
        # Bind mouse wheel to spinboxes
        self.x_offset_spinbox.bind('<MouseWheel>', self.handle_mouse_wheel)
        self.y_offset_spinbox.bind('<MouseWheel>', self.handle_mouse_wheel)
        self.opacity_spinbox.bind('<MouseWheel>', self.handle_mouse_wheel)
        
        # Status Label
        self.status_label = ttk.Label(self.main_frame, text="Ready", justify=tk.CENTER)
        self.status_label.pack(pady=5)
        
    def handle_mouse_wheel(self, event):
        """Handle mouse wheel events for the spinboxes."""
        widget = event.widget
        try:
            current_value = int(widget.get())
            if event.delta > 0:
                widget.set(str(current_value + 1))
            else:
                if current_value > 1:  # Prevent going below 1
                    widget.set(str(current_value - 1))
        except ValueError:
            # If the value can't be converted to int, do nothing
            pass
            
    def setup_drag_drop(self):
        if not HAS_TKINTERDND:
            return
            
        def handle_dir_drop(event):
            if not event.data:
                return
                
            # Clean up the dropped paths
            paths = [path.strip('{}') for path in event.data.split()]
            valid_paths = []
            
            for path in paths:
                if os.path.exists(path):
                    valid_paths.append(path)
                else:
                    # Try to handle Windows paths with backslashes
                    path = path.replace('\\', '/')
                    if os.path.exists(path):
                        valid_paths.append(path)
            
            if valid_paths:
                # If multiple files were dropped, use the directory containing the first file
                if len(valid_paths) > 1:
                    self.input_dir.set(os.path.dirname(valid_paths[0]))
                else:
                    # If a single file was dropped, use its directory
                    self.input_dir.set(os.path.dirname(valid_paths[0]))
                
                # Update output directory based on input directory
                input_path = Path(self.input_dir.get())
                self.output_dir.set(str(input_path.parent / f"BATES_{input_path.name}"))
                
                # Update source label
                display_path = str(input_path)
                if len(display_path) > 50:
                    display_path = "..." + display_path[-47:]
                self.source_label.config(text=display_path)
                
                # Enable process button
                self.stamp_dir_button.config(state=tk.NORMAL)
        
        def handle_dest_drop(event):
            if not event.data:
                return
                
            # Clean up the dropped paths
            paths = [path.strip('{}') for path in event.data.split()]
            valid_paths = []
            
            for path in paths:
                if os.path.exists(path):
                    valid_paths.append(path)
                else:
                    # Try to handle Windows paths with backslashes
                    path = path.replace('\\', '/')
                    if os.path.exists(path):
                        valid_paths.append(path)
            
            if valid_paths:
                # Use the first valid path
                dest_path = valid_paths[0]
                self.output_dir.set(dest_path)
                
                # Update destination label
                display_path = str(dest_path)
                if len(display_path) > 50:
                    display_path = "..." + display_path[-47:]
                self.dest_label.config(text=display_path)
        
        def handle_file_drop(event):
            if not event.data:
                return
                
            # Clean up the dropped paths
            paths = [path.strip('{}') for path in event.data.split()]
            valid_paths = []
            
            for path in paths:
                if os.path.exists(path):
                    valid_paths.append(path)
                else:
                    # Try to handle Windows paths with backslashes
                    path = path.replace('\\', '/')
                    if os.path.exists(path):
                        valid_paths.append(path)
            
            if valid_paths:
                # Use the first valid file
                file_path = valid_paths[0]
                self.input_file.set(file_path)
                
                # Set default output directory to same location as input file
                self.file_output_dir.set(str(Path(file_path).parent))
                
                # Update file label
                self.file_label.config(text=Path(file_path).name)
                
                # Enable process button
                self.stamp_file_button.config(state=tk.NORMAL)
        
        def handle_file_dest_drop(event):
            if not event.data:
                return
                
            # Clean up the dropped paths
            paths = [path.strip('{}') for path in event.data.split()]
            valid_paths = []
            
            for path in paths:
                if os.path.exists(path):
                    valid_paths.append(path)
                else:
                    # Try to handle Windows paths with backslashes
                    path = path.replace('\\', '/')
                    if os.path.exists(path):
                        valid_paths.append(path)
            
            if valid_paths:
                # Use the first valid path
                dest_path = valid_paths[0]
                self.file_output_dir.set(dest_path)
                
                # Update destination label
                display_path = str(dest_path)
                if len(display_path) > 50:
                    display_path = "..." + display_path[-47:]
                self.file_dest_label.config(text=display_path)
        
        # Make the source directory drop zone a drop target
        self.source_label.drop_target_register(DND_FILES)
        self.source_label.dnd_bind('<<Drop>>', handle_dir_drop)
        
        # Make the destination directory drop zone a drop target
        self.dest_label.drop_target_register(DND_FILES)
        self.dest_label.dnd_bind('<<Drop>>', handle_dest_drop)
        
        # Make the file drop zone a drop target
        self.file_label.drop_target_register(DND_FILES)
        self.file_label.dnd_bind('<<Drop>>', handle_file_drop)
        
        # Make the file destination drop zone a drop target
        self.file_dest_label.drop_target_register(DND_FILES)
        self.file_dest_label.dnd_bind('<<Drop>>', handle_file_dest_drop)
        
    def browse_folder(self):
        """Open folder browser dialog."""
        folder_path = filedialog.askdirectory(
            title="Select Source Folder",
            initialdir=self.last_directory
        )
        if folder_path:
            path = Path(folder_path)
            self.input_dir.set(str(path))
            # Update last directory
            self.last_directory = str(path)
            # Set default output directory to parent directory with BATES_ prefix
            self.output_dir.set(str(path.parent / f"BATES_{path.name}"))
            # Truncate long paths for display
            display_path = str(path)
            if len(display_path) > 50:
                display_path = "..." + display_path[-47:]
            self.source_label.config(text=display_path)
            
    def browse_output_folder(self):
        """Open folder browser dialog for output directory."""
        folder_path = filedialog.askdirectory(
            title="Select Destination Folder",
            initialdir=self.last_directory
        )
        if folder_path:
            self.output_dir.set(folder_path)
            # Update last directory
            self.last_directory = folder_path
            
    def browse_file(self):
        """Open file browser dialog."""
        file_path = filedialog.askopenfilename(
            title="Select File",
            initialdir=self.last_directory,
            filetypes=[
                ("PDF files", ".pdf"),
                ("Excel files", (".xlsx", ".xls", ".xlsm")),
                ("Image files", (".png", ".jpg", ".jpeg", ".gif", ".bmp")),
                ("All files", "*.*")
            ]
        )
        if file_path:
            path = Path(file_path)
            self.input_file.set(str(path))
            self.file_label.config(text=path.name)
            # Update last directory
            self.last_directory = str(path.parent)
            # Set default output directory to same location as input file
            self.file_output_dir.set(str(path.parent))
            # Reset status message when new file is selected
            self.status_label.config(text="Ready")
            
    def browse_file_output_folder(self):
        """Open folder browser dialog for file output directory."""
        folder_path = filedialog.askdirectory(
            title="Select Destination Folder",
            initialdir=self.last_directory
        )
        if folder_path:
            self.file_output_dir.set(folder_path)
            # Update last directory
            self.last_directory = folder_path
            
    def start_processing(self):
        if not self.input_dir.get():
            messagebox.showerror("Error", "Please select a source directory")
            return
            
        if self.processing:
            return
            
        self.processing = True
        self.stamp_dir_button.state(['disabled'])
        self.status_label.config(text="Processing...")
        
        try:
            # Clean up prefix (remove trailing underscore if present)
            prefix = self.prefix.get().rstrip('_')
            
            # Create output directory with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_dir = Path(self.output_dir.get()) / f"BATES_{prefix}_{timestamp}"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            processor = EnhancedBatesNumbering(
                input_dir=self.input_dir.get(),
                output_dir=str(output_dir),  # Use the new timestamped directory
                prefix=prefix,
                zero_pad_length=self.digits.get(),
                start=self.start_number.get(),
                stamp_x=self.stamp_x.get(),
                stamp_y=self.stamp_y.get(),
                stamp_color=self.stamp_color.get(),
                stamp_box_width=self.stamp_box_width.get(),
                stamp_position=self.stamp_position.get(),
                stamp_x_offset=self.stamp_x_offset.get(),
                stamp_y_offset=self.stamp_y_offset.get(),
                stamp_opacity=self.stamp_opacity.get()
            )
            processor.run()
            messagebox.showinfo("Success", "Processing completed successfully!")
            
            # Open output folder
            self.open_output_folder(str(output_dir))
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            self.processing = False
            self.stamp_dir_button.state(['!disabled'])
            self.status_label.config(text="Ready")
            
    def stamp_single_file(self):
        """Process a single file for Bates stamping."""
        if not self.input_file.get():
            self.status_label.config(text="Please select a file")
            return
            
        if self.processing:
            return
            
        self.processing = True
        self.stamp_file_button.state(['disabled'])
        self.status_label.config(text="Processing...")
        
        try:
            input_file = Path(self.input_file.get())
            output_dir = Path(self.file_output_dir.get())
            
            # For single file, don't create timestamped directory
            processor = EnhancedBatesNumbering(
                input_dir=str(input_file.parent),
                output_dir=str(output_dir),
                prefix=self.prefix.get().rstrip('_'),
                zero_pad_length=self.digits.get(),
                start=self.start_number.get(),
                is_single_file=True,  # Set this to True for single file processing
                stamp_x=self.stamp_x.get(),
                stamp_y=self.stamp_y.get(),
                stamp_color=self.stamp_color.get(),
                stamp_box_width=self.stamp_box_width.get(),
                stamp_position=self.stamp_position.get(),
                stamp_x_offset=self.stamp_x_offset.get(),
                stamp_y_offset=self.stamp_y_offset.get(),
                stamp_opacity=self.stamp_opacity.get()
            )
            
            # Process just this file
            if input_file.suffix.lower() == '.pdf':
                bates_number = f"{processor.prefix}{str(processor.current_number).zfill(processor.zero_pad_length)}"
                output_pdf = output_dir / f"{bates_number}_{input_file.name}"
                if processor.add_bates_stamp(input_file, output_pdf, bates_number):
                    self.status_label.config(text=f"File stamped successfully: {output_pdf.name}")
                    self.open_output_folder(str(output_dir))
                else:
                    self.status_label.config(text="File was copied to _FILES WITH ISSUES folder due to stamping issues")
            else:
                # Convert to PDF first
                pdf_path = processor.convert_to_pdf(input_file, output_dir)
                if pdf_path:
                    bates_number = f"{processor.prefix}{str(processor.current_number).zfill(processor.zero_pad_length)}"
                    output_pdf = output_dir / f"{bates_number}_{pdf_path.name}"
                    if processor.add_bates_stamp(pdf_path, output_pdf, bates_number):
                        self.status_label.config(text=f"File converted and stamped successfully: {output_pdf.name}")
                        self.open_output_folder(str(output_dir))
                    else:
                        self.status_label.config(text="File was copied to _FILES WITH ISSUES folder due to stamping issues")
                else:
                    self.status_label.config(text="Could not convert file to PDF")
            
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}")
        finally:
            self.processing = False
            self.stamp_file_button.state(['!disabled'])
            
    def open_output_folder(self, custom_output_dir=None):
        """Open the output folder in the system's file explorer."""
        output_dir = custom_output_dir if custom_output_dir else self.output_dir.get()
        if output_dir:
            if sys.platform == 'win32':
                os.startfile(output_dir)
            elif sys.platform == 'darwin':
                subprocess.run(['open', output_dir])
            else:
                subprocess.run(['xdg-open', output_dir])
    
    def run(self):
        self.root.mainloop()

def main():
    if len(sys.argv) > 1:
        # Command line mode
        parser = argparse.ArgumentParser(description='Enhanced Bates Numbering Utility')
        parser.add_argument('input_path', help='Input file or directory')
        parser.add_argument('output_dir', help='Directory for processed files')
        parser.add_argument('--prefix', default='', help='Prefix for Bates numbers')
        parser.add_argument('--zero-pad', type=int, default=5, 
                          help='Number of zeros to pad Bates numbers (default: 5)')
        parser.add_argument('--start', type=int, default=1, 
                          help='Starting number for Bates numbering')
        
        args = parser.parse_args()
        
        # Clean up prefix (remove trailing underscore if present)
        prefix = args.prefix.rstrip('_')
        
        processor = EnhancedBatesNumbering(
            input_dir=str(Path(args.input_path).parent),
            output_dir=args.output_dir,
            prefix=prefix,
            zero_pad_length=args.zero_pad,
            start=args.start
        )
        
        # Check if input is a single file
        input_path = Path(args.input_path)
        if input_path.is_file():
            if input_path.suffix.lower() == '.pdf':
                bates_number = f"{processor.prefix}{str(processor.current_number).zfill(processor.zero_pad_length)}"
                output_pdf = Path(args.output_dir) / f"{bates_number}_{input_path.name}"
                if processor.add_bates_stamp(input_path, output_pdf, bates_number):
                    print(f"File stamped successfully: {output_pdf}")
                else:
                    print("File was copied to BATES - LOCKED folder due to stamping issues")
            else:
                # Convert to PDF first
                pdf_path = processor.convert_to_pdf(input_path, Path(args.output_dir))
                if pdf_path:
                    bates_number = f"{processor.prefix}{str(processor.current_number).zfill(processor.zero_pad_length)}"
                    output_pdf = Path(args.output_dir) / f"{bates_number}_{pdf_path.name}"
                    if processor.add_bates_stamp(pdf_path, output_pdf, bates_number):
                        print(f"File converted and stamped successfully: {output_pdf}")
                    else:
                        print("File was copied to BATES - LOCKED folder due to stamping issues")
                else:
                    print("Could not convert file to PDF")
        else:
            # Process entire directory
            processor.run()
    else:
        # GUI mode
        gui = BatesGUI()
        gui.run()

if __name__ == '__main__':
    main() 